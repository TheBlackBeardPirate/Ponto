import calendar
import datetime
import flet
from flet import *
import openpyxl

CELL_SIZE = (28, 28)
CELL_BG_COLOR = 'white10'
TODAY_BG_COLOR = 'teal600'


class SetCalendar(UserControl):
    def __init__(self, start_year=datetime.date.today().year):

        self.current_year = start_year
        self.m1 = datetime.date.today().month
        self.m2 = self.m1 + 1

        self.click_count: list = []
        self.long_press_count: list = []

        self.current_color = 'blue'
        self.selected_date = any
        self.selected_date = None

        # Carga horária (editei aqui)
        self.date_count = []
        self.carga_horaria = any

        self.calendar_grid = Column(
            wrap=True,
            alignment=MainAxisAlignment.CENTER,
            horizontal_alignment=CrossAxisAlignment.CENTER,
        )
        super().__init__()

    def _change_month(self, delta):
        self.m1 = min(max(1, self.m1 + delta), 12)
        self.m2 = min(max(2, self.m2 + delta), 13)

        new_calendar = self.create_month_calendar(self.current_year)
        self.calendar_grid = new_calendar
        self.update()

    def one_click_date(self, e):
        if e.control.data not in self.date_count:
            self.selected_date = e.control.data
            e.control.bgcolor = 'blue600'
            e.control.update()
            self.update()
            self.date_count.append(self.selected_date)
            # print(self.date_count)
        else:
            e.control.bgcolor = colors.SURFACE_VARIANT,
            e.control.update()
            n = 0
            for x in self.date_count:
                if x != e.control.data:
                    n += 1
                else:
                    self.date_count.pop(n)
            self.update()
            # print(self.date_count)

    def long_click_date(self, e):
        self.long_press_count.append(e.control.data)
        if len(self.long_press_count) == 2:
            date1, date2 = self.long_press_count
            delta = abs(date2 - date1)
            if date1 < date2:
                dates = [
                    date1 + datetime.timedelta(days=x) for x in range(delta.days + 1)
                ]
            else:
                dates = [
                    date2 + datetime.timedelta(days=x) for x in range(delta.days + 1)
                ]

            for _ in self.calendar_grid.controls[:]:
                for __ in _.controls[:]:
                    if isinstance(__, Row):
                        for box in __.controls[:]:
                            if box.data in dates:
                                box.bgcolor = 'blue600'
                                box.update()

            self.long_press_count = []
        else:
            pass

    def create_month_calendar(self, year):
        self.current_year = year
        self.calendar_grid.controls: list = []

        for month in range(self.m1, self.m2):
            month_label = Text(
                value=f'{calendar.month_name[month]} {self.current_year}',
                color='white',
                size=14,
                weight='bold',
            )

            month_matrix = calendar.monthcalendar(self.current_year, month)
            month_grid = Column(alignment=MainAxisAlignment.CENTER)
            month_grid.controls.append(
                Row(
                    alignment=MainAxisAlignment.START,
                    controls=[month_label],
                )
            )
            weekday_labels = [
                Container(
                    width=28,
                    height=28,
                    alignment=alignment.center,
                    content=Text(
                        weekday,
                        size=12,
                        color='white54',
                    )
                )
                for weekday in ['Seg', 'Ter', 'Qua', 'Qui', 'Sex', 'Sáb', 'Dom']
            ]

            weekday_row = Row(controls=weekday_labels)
            month_grid.controls.append(weekday_row)

            for week in month_matrix:
                week_container = Row()
                for day in week:
                    if day == 0:
                        day_container = Container(
                            width=28,
                            height=28,
                        )
                    else:
                        day_container = Container(
                            width=28,
                            height=28,
                            border=border.all(0.5, 'white54'),
                            alignment=alignment.center,
                            data=datetime.date(
                                year=self.current_year,
                                month=month,
                                day=day
                            ),
                            on_click=lambda e: self.one_click_date(e),
                            on_long_press=lambda e: self.long_click_date(e),
                            animate=400,
                        )
                    day_label = Text(str(day), size=12)

                    if day == 0:
                        day_label = None
                    if (
                            day == datetime.date.today().day
                            and month == datetime.date.today().month
                            and self.current_year == datetime.date.today().year
                    ):
                        day_container.bgcolor = 'teal700'
                    day_container.content = day_label
                    week_container.controls.append(day_container)
                month_grid.controls.append(week_container)

        self.calendar_grid.controls.append(month_grid)

        return self.calendar_grid

    def date(self):
        return self.date_count

    def build(self):
        return self.create_month_calendar(self.current_year)


class DateSetUp(UserControl):
    def __init__(self, cal_grid):
        self.cal_grid = cal_grid

        self.prev_btn = BTNPagination('Prev.', lambda e: cal_grid._change_month(-1))
        self.next_btn = BTNPagination('Next', lambda e: cal_grid._change_month(1))

        self.today = Text(
            datetime.date.today().strftime('%B %d, %Y'),
            width=260,
            size=13,
            color='white54',
            weight='w400'
        )

        self.btn_container = Row(
            alignment=MainAxisAlignment.CENTER,
            controls=[
                self.prev_btn,
                self.next_btn
            ]
        )

        self.calendar = Container(
            width=320,
            height=45,
            bgcolor='#313131',
            border_radius=8,
            animate=300,
            clip_behavior=ClipBehavior.HARD_EDGE,
            alignment=alignment.center,
            content=Column(
                alignment=MainAxisAlignment.START,
                horizontal_alignment=CrossAxisAlignment.CENTER,
                controls=[
                    Divider(height=60, color='transparent'),
                    self.cal_grid,
                    Divider(height=60, color='transparent'),
                    self.btn_container,
                ]
            )
        )

        super().__init__()

    def _get_calendar(self, e: None):
        if self.calendar.height == 45:
            self.calendar.height = 450
            self.calendar.update()
        else:
            self.calendar.height = 45
            self.calendar.update()

    def build(self):
        return Stack(
            width=320,
            controls=[
                self.calendar,
                Container(
                    on_click=lambda e: self._get_calendar(e),
                    width=320,
                    height=45,
                    border_radius=8,
                    bgcolor='#313131',
                    padding=padding.only(left=15, right=5),
                    content=Row(
                        alignment=MainAxisAlignment.SPACE_BETWEEN,
                        vertical_alignment=CrossAxisAlignment.CENTER,
                        controls=[
                            self.today,
                            Container(
                                width=32,
                                height=32,
                                border=border.only(
                                    left=BorderSide(0.9, 'white24'),
                                ),
                                alignment=alignment.center,
                                content=Icon(
                                    name=icons.CALENDAR_MONTH_SHARP,
                                    size=15,
                                    opacity=0.65
                                )
                            )
                        ],
                    )
                )
            ]
        )


class BTNPagination(UserControl):
    def __init__(self, txt_name, function):
        self.txt_name = txt_name
        self.function = function
        super().__init__()

    def build(self):
        return IconButton(
            content=Text(self.txt_name, size=8, weight='bold'),
            width=56,
            height=28,
            on_click=self.function,
            style=ButtonStyle(
                shape={
                    '': RoundedRectangleBorder(radius=6)
                },
                bgcolor={
                    '': 'teal600'
                }
            )
        )


def calc_horas_mensais(horario, multiplicador):
    minutos_originais = int(horario.split(':')[0]) * 60 + int(horario.split(':')[1])

    minutos_resultado = minutos_originais * multiplicador

    horas, minutos = divmod(minutos_resultado, 60)

    resultado_formatado = f'{horas:02d}:{minutos:02d}'

    return resultado_formatado


def calc_horas_diarias(hora_entrada_manha, hora_saida_manha, hora_entrada_tarde, hora_saida_tarde):
    horario1 = datetime.datetime.strptime(hora_entrada_manha, '%H:%M')
    horario2 = datetime.datetime.strptime(hora_saida_manha, '%H:%M')
    horario3 = datetime.datetime.strptime(hora_entrada_tarde, '%H:%M')
    horario4 = datetime.datetime.strptime(hora_saida_tarde, '%H:%M')

    diferenca = (horario2 - horario1) + (horario4 - horario3)

    horas, minutos = divmod(diferenca.seconds, 3600)
    minutos //= 60

    horas_trabalhadas = f'{horas:02d}:{minutos:02d}'

    return horas_trabalhadas


def calc_hora_extra(hora_extra_entrada, hora_extra_saida):
    horario1 = datetime.datetime.strptime(hora_extra_entrada, '%H:%M')
    horario2 = datetime.datetime.strptime(hora_extra_saida, '%H:%M')

    diferenca = horario2 - horario1

    horas, minutos = divmod(diferenca.seconds, 3600)
    minutos //= 60

    horas_trabalhadas = f'{horas:02d}:{minutos:02d}'

    return horas_trabalhadas


def lista_dias(data):
    primeiro_dia_do_mes = datetime.datetime(data.year, data.month, 1)
    ultimo_dia_do_mes = datetime.datetime(data.year, data.month, calendar.monthrange(data.year, data.month)[1])
    dias_do_mes = [primeiro_dia_do_mes + datetime.timedelta(days=d) for d in
                   range((ultimo_dia_do_mes - primeiro_dia_do_mes).days + 1)]

    lista = []

    for dia in dias_do_mes:
        formato_data = dia.date()
        dia_da_semana_abreviado = dia.strftime("%a")

        if dia_da_semana_abreviado == 'Mon':
            dia_da_semana_abreviado = 'Seg'
        elif dia_da_semana_abreviado == 'Tue':
            dia_da_semana_abreviado = 'Ter'
        elif dia_da_semana_abreviado == 'Wed':
            dia_da_semana_abreviado = 'Qua'
        elif dia_da_semana_abreviado == 'Thu':
            dia_da_semana_abreviado = 'Qui'
        elif dia_da_semana_abreviado == 'Fri':
            dia_da_semana_abreviado = 'Sex'
        elif dia_da_semana_abreviado == 'Sat':
            dia_da_semana_abreviado = 'Sáb'
        elif dia_da_semana_abreviado == 'Sun':
            dia_da_semana_abreviado = 'Dom'

        dicionario = {'data': formato_data, 'dia_da_semana_abreviado': dia_da_semana_abreviado}
        lista.append(dicionario)

    return lista


def ponto(funcionario, empresa, expediente_normal, hora_entrada_manha, hora_saida_manha, hora_entrada_tarde,
          hora_saida_tarde, hora_extra, hora_extra_entrada, hora_extra_saida):
    workbook = openpyxl.load_workbook('calcular_horas.xlsx')
    sheet = workbook['Planilha1']
    sheet.cell(row=1, column=4, value=empresa)
    sheet.cell(row=2, column=4, value=funcionario)
    sheet.cell(row=2, column=14, value=expediente_normal[0].strftime("%m/%Y"))
    lista_mes = lista_dias(expediente_normal[0])
    i = 4
    j = 0
    n = 0
    tam1 = len(expediente_normal)
    tam2 = len(hora_extra)
    horas_trabalhadas = ''

    diferenca = calc_hora_extra(hora_extra_entrada, hora_extra_saida)

    for data in lista_mes:
        sheet.cell(row=i, column=1, value=data['data'])
        sheet.cell(row=i, column=2, value=data['dia_da_semana_abreviado'])

        if j < tam1 and expediente_normal[j] == data['data']:
            sheet.cell(row=i, column=4, value=hora_entrada_manha)
            sheet.cell(row=i, column=5, value=hora_saida_manha)
            sheet.cell(row=i, column=6, value=hora_entrada_tarde)
            sheet.cell(row=i, column=7, value=hora_saida_tarde)

            horas_trabalhadas = calc_horas_diarias(hora_entrada_manha, hora_saida_manha, hora_entrada_tarde, hora_saida_tarde)

            sheet.cell(row=i, column=10, value=horas_trabalhadas)
            j += 1

        if n < tam2 and hora_extra[n] == data['data']:
            sheet.cell(row=i, column=11, value=diferenca)
            n += 1

        i += 1

    horas_trabalhadas = calc_horas_mensais(horas_trabalhadas, j)
    sheet.cell(row=35, column=10, value=horas_trabalhadas)
    horas_trabalhadas = calc_horas_mensais(diferenca, n)
    sheet.cell(row=35, column=11, value=horas_trabalhadas)

    workbook.save('calcular_horas.xlsx')


def main(page: Page):
    page.title = 'Ponto Eletrônico'
    page.theme_mode = ThemeMode.DARK

    page.horizontal_alignments = 'center'
    page.vertical_alignments = 'center'
    page.padding = 80
    page.update()

    cal1 = SetCalendar()
    cal2 = SetCalendar()
    date_normal = DateSetUp(cal1)
    date_hora_extra = DateSetUp(cal2)

    empresas = [dropdown.Option('Enge10'), dropdown.Option('Pré-Moldados'), dropdown.Option('Comercial Inova'),
                dropdown.Option('Outra')]

    dropdown_empresa = Dropdown(
        width=310,
        label='Empresa',
        hint_text="Empresa",
        options=empresas,
    )

    funcionario = TextField(label="Nome do Funcionário", hint_text="Nome do Funcionário", width=310)
    hora_entrada_manha = TextField(label="Horário de entrada de manhã", hint_text="7:30",
                                   width=310)
    hora_saida_manha = TextField(label="Horário de saída de manhã", hint_text="11:30", width=310)
    hora_entrada_tarde = TextField(label="Horário de entrada de tarde", hint_text="13:30",
                                   width=310)
    hora_saida_tarde = TextField(label="Horário de saída de tarde", hint_text="17:30", width=310)

    hora_extra_text = Text(value="Hora Extra", size=28)

    hora_extra_entrada = TextField(label="Horário de início", hint_text="17:30", width=310)

    hora_extra_saida = TextField(label="Horário de fim", hint_text="23:00", width=310)

    def mudanca_de_rota(route):
        def func(e):
            expediente_normal = date_normal.cal_grid.date()
            hora_extra = date_hora_extra.cal_grid.date()

            ponto(funcionario.value, dropdown_empresa.value, expediente_normal, hora_entrada_manha.value,
                  hora_saida_manha.value, hora_entrada_tarde.value, hora_saida_tarde.value, hora_extra,
                  hora_extra_entrada.value, hora_extra_saida.value)

        page.views.clear()

        page.views.append(
            View(
                route='/',
                controls=[
                    AppBar(
                        title=Text('Ponto Eletrônico'),
                        bgcolor=colors.SURFACE_VARIANT,
                        actions=[],
                    ),
                    Column(
                        scroll=ScrollMode(value='hidden'),
                        expand=True,
                        controls=[
                            Row([
                                Divider(height=9, thickness=3),
                                dropdown_empresa,
                                funcionario,
                            ]),
                            Row([
                                Divider(height=9, thickness=3),
                                hora_entrada_manha,
                                hora_saida_manha,
                            ]),
                            Row([
                                Divider(height=9, thickness=3),
                                hora_entrada_tarde,
                                hora_saida_tarde,
                            ]),
                            Row([
                                Divider(height=9, thickness=3),
                                date_normal,
                            ]),
                            Row([
                                Divider(height=9, thickness=3),
                                hora_extra_text
                            ]),
                            Row([
                                Divider(height=9, thickness=3),
                                date_hora_extra
                            ]),
                            Row([
                                Divider(height=9, thickness=3),
                                hora_extra_entrada,
                                hora_extra_saida
                            ]),
                            Row([
                                Divider(height=9, thickness=3),
                                ElevatedButton(text='Gerar Ficha Ponto', color='#00C60C', icon="add", bgcolor='#1445A6',
                                               height=50, on_click=func)
                            ])
                        ],
                    )
                ]
            )
        )
        if page.route == '/m':
            pass

        page.update()

    def view_pop():
        page.views.pop()
        top_view = page.views[-1]
        page.go(top_view.route)

    page.on_route_change = mudanca_de_rota
    page.on_route_change = view_pop
    page.go(page.route)


if __name__ == '__main__':
    flet.app(target=main)
